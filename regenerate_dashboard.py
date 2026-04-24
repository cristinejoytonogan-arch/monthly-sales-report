"""
KSESTOR Monthly Sales Report - HTML Dashboard Regenerator
=========================================================

Reads your 'Monthly Report - Clean Version v2.xlsx' and rebuilds the HTML dashboard
'Monthly Report - Dashboard v2.html' in the same folder.

USAGE:
    1. Update the Excel file (Sales Data, Main Expenses, Sub Expenses sheets).
    2. Save the Excel file.
    3. Open Command Prompt / PowerShell in this folder.
    4. Run: python regenerate_dashboard.py
    5. The HTML file will be overwritten with the latest data.

REQUIREMENTS:
    - Python 3.10 or newer
    - openpyxl: install with `pip install openpyxl`

Prepared for Kseniia · KSESTOR
"""

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Please run:  pip install openpyxl")
    sys.exit(1)

# Resolve paths relative to this script
SCRIPT_DIR = Path(__file__).parent
EXCEL_FILE = SCRIPT_DIR / "Monthly Report - Clean Version v2.xlsx"
HTML_FILE = SCRIPT_DIR / "index.html"  # outputs directly as index.html for GitHub Pages

if not EXCEL_FILE.exists():
    print(f"ERROR: Cannot find {EXCEL_FILE.name}")
    print(f"Expected it at: {EXCEL_FILE}")
    sys.exit(1)

print(f"Reading {EXCEL_FILE.name}...")

# ===== Load the Excel data =====
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=False)

# -- Sales Data --
ws = wb["Sales Data"]
sales = []
# Header row is row 5, data starts row 6
for r in range(6, ws.max_row + 1):
    dt = ws.cell(row=r, column=1).value
    mp = ws.cell(row=r, column=2).value
    if not dt or not mp:
        continue
    # Month & year are in cols 55 (BC) and 56 (BD)
    m = ws.cell(row=r, column=55).value
    y = ws.cell(row=r, column=56).value
    sales.append({
        "date": dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt),
        "month": m, "year": y, "mp": mp,
        "main_cat": ws.cell(row=r, column=3).value,
        "sub_cat": ws.cell(row=r, column=4).value,
        "sales": ws.cell(row=r, column=5).value or 0,
        "units": ws.cell(row=r, column=7).value,
        "vat": ws.cell(row=r, column=10).value,
        "net_profit": ws.cell(row=r, column=12).value,
        "tacos": ws.cell(row=r, column=13).value,
        "refunds": ws.cell(row=r, column=14).value,
        "sellable_returns": ws.cell(row=r, column=15).value,
        "margin": ws.cell(row=r, column=16).value,
        "roi": ws.cell(row=r, column=17).value,
        # TikTok cols
        "tt_net_sales": ws.cell(row=r, column=18).value,
        "tt_shipping": ws.cell(row=r, column=19).value,
        "tt_3pl": ws.cell(row=r, column=20).value,
        "tt_mcf": ws.cell(row=r, column=21).value,
        "tt_fees": ws.cell(row=r, column=22).value,
        "tt_referral": ws.cell(row=r, column=23).value,
        "tt_affiliate": ws.cell(row=r, column=24).value,
        "tt_adj": ws.cell(row=r, column=25).value,
        "tt_cogs": ws.cell(row=r, column=26).value,
        "tt_ads": ws.cell(row=r, column=27).value,
        "tt_np": ws.cell(row=r, column=28).value,
        "tt_margin": ws.cell(row=r, column=29).value,
        # Shopify cols
        "sh_ads": ws.cell(row=r, column=30).value,
        "sh_ship": ws.cell(row=r, column=31).value,
        "sh_mcf": ws.cell(row=r, column=32).value,
        "sh_3pl": ws.cell(row=r, column=33).value,
        "sh_pay": ws.cell(row=r, column=34).value,
        "sh_cogs": ws.cell(row=r, column=35).value,
        "sh_tax": ws.cell(row=r, column=36).value,
        "sh_gp": ws.cell(row=r, column=37).value,
        "sh_np": ws.cell(row=r, column=38).value,
        "sh_tacos": ws.cell(row=r, column=39).value,
        "sh_refunds": ws.cell(row=r, column=40).value,
        "sh_margin": ws.cell(row=r, column=41).value,
        "sh_roi": ws.cell(row=r, column=42).value,
        # Walmart cols
        "wm_ads": ws.cell(row=r, column=43).value,
        "wm_wfs": ws.cell(row=r, column=44).value,
        "wm_mcf": ws.cell(row=r, column=45).value,
        "wm_3pl": ws.cell(row=r, column=46).value,
        "wm_fee_com": ws.cell(row=r, column=47).value,
        "wm_fee_store": ws.cell(row=r, column=48).value,
        "wm_other": ws.cell(row=r, column=49).value,
        "wm_cogs": ws.cell(row=r, column=50).value,
        "wm_np": ws.cell(row=r, column=51).value,
    })
print(f"  Sales Data: {len(sales)} rows")

# -- Main Expenses --
ws = wb["Main Expenses"]
expenses = []
for r in range(6, ws.max_row + 1):
    dt = ws.cell(row=r, column=1).value
    mp = ws.cell(row=r, column=2).value
    if not dt or not mp:
        continue
    expenses.append({
        "date": dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt),
        "month": ws.cell(row=r, column=11).value,
        "year": ws.cell(row=r, column=12).value,
        "mp": mp,
        "category": ws.cell(row=r, column=4).value,
        "amount": ws.cell(row=r, column=5).value or 0,
    })
print(f"  Main Expenses: {len(expenses)} rows")

# -- Sub Expenses --
ws = wb["Sub Expenses"]
sub_expenses = []
for r in range(6, ws.max_row + 1):
    dt = ws.cell(row=r, column=1).value
    mp = ws.cell(row=r, column=2).value
    main = ws.cell(row=r, column=3).value
    sub = ws.cell(row=r, column=4).value
    amt = ws.cell(row=r, column=5).value
    m = ws.cell(row=r, column=10).value
    y = ws.cell(row=r, column=11).value
    if not (mp and main and sub):
        continue
    sub_expenses.append({
        "date": dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt),
        "month": m, "year": y, "mp": mp,
        "main": main.strip() if isinstance(main, str) else main,
        "sub": sub,
        "amount": amt or 0,
    })
print(f"  Sub Expenses: {len(sub_expenses)} rows")

wb.close()

# ===== Build the HTML =====
from datetime import date as _date
data = {
    "sales": sales,
    "expenses": expenses,
    "sub_expenses": sub_expenses,
    "generated": _date.today().strftime("%B %d, %Y"),
}
data_js = json.dumps(data, default=str)

HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>KSESTOR · Monthly Sales Report Dashboard</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root {
  --navy: #272757; --navy-dark: #1C1D3D; --navy-light: #E5E7F3; --soft-tint: #EFF0F9;
  --black: #0F172A; --slate-50: #F8FAFC; --slate-100: #F1F5F9; --slate-200: #E2E8F0;
  --slate-300: #CBD5E1; --slate-500: #64748B; --slate-600: #475569; --slate-700: #334155;
  --pos: #059669; --neg: #DC2626;
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #FFFFFF; color: var(--black); line-height: 1.5; -webkit-font-smoothing: antialiased; font-feature-settings: 'tnum' on, 'lnum' on; }
.container { max-width: 1440px; margin: 0 auto; padding: 40px 32px; }
.header { border-bottom: 3px solid var(--navy); padding-bottom: 20px; margin-bottom: 32px; }
.header h1 { font-size: 32px; font-weight: 800; letter-spacing: -0.03em; color: var(--black); }
.header h2 { font-size: 13px; color: var(--navy); text-transform: uppercase; letter-spacing: 0.12em; margin-top: 6px; font-weight: 700; }
.header .meta { color: var(--slate-500); font-size: 12px; margin-top: 8px; font-style: italic; }
.tabs { display: flex; gap: 4px; margin-bottom: 24px; border-bottom: 1px solid var(--slate-200); overflow-x: auto; }
.tab { padding: 12px 20px; border: none; background: transparent; color: var(--slate-500); font-weight: 600; font-size: 13px; cursor: pointer; border-bottom: 3px solid transparent; white-space: nowrap; transition: all 0.15s; font-family: inherit; letter-spacing: 0.01em; }
.tab:hover { color: var(--black); }
.tab.active { color: var(--navy); border-bottom-color: var(--navy); }
.tab-content { display: none; }
.tab-content.active { display: block; }
.selectors { display: flex; gap: 10px; margin-bottom: 24px; flex-wrap: wrap; align-items: flex-end; }
.sel-group { display: flex; flex-direction: column; gap: 6px; }
.sel-group label { font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; color: var(--slate-500); }
.sel-group select { padding: 9px 14px; border: 2px solid var(--navy); border-radius: 5px; background: var(--navy-light); font-weight: 700; font-size: 13px; cursor: pointer; min-width: 130px; font-family: inherit; color: var(--navy-dark); }
.sel-group select:focus { outline: 2px solid var(--navy); outline-offset: 1px; }
.kpi-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 10px; margin-bottom: 28px; }
.kpi { background: #FFF; border: 1px solid var(--slate-200); border-left: 3px solid var(--navy); padding: 14px 16px; border-radius: 4px; }
.kpi .label { font-size: 10px; text-transform: uppercase; letter-spacing: 0.1em; color: var(--slate-500); font-weight: 700; margin-bottom: 4px; }
.kpi .value { font-size: 18px; font-weight: 700; letter-spacing: -0.02em; color: var(--navy-dark); font-variant-numeric: tabular-nums; }
.kpi .value.pos { color: var(--pos); } .kpi .value.neg { color: var(--neg); }
.section { margin-bottom: 28px; }
.section-head { display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; padding-bottom: 8px; border-bottom: 2px solid var(--black); }
.section-head h3 { font-size: 13px; text-transform: uppercase; letter-spacing: 0.1em; font-weight: 700; color: var(--black); }
.section-head .inline-sel { display: flex; gap: 8px; align-items: center; }
.section-head .inline-sel label { font-size: 10px; text-transform: uppercase; letter-spacing: 0.08em; color: var(--slate-500); font-weight: 700; }
.section-head .inline-sel select { padding: 5px 10px; border: 1.5px solid var(--navy); border-radius: 4px; background: var(--navy-light); font-weight: 600; font-size: 12px; color: var(--navy-dark); cursor: pointer; font-family: inherit; }
table { width: 100%; border-collapse: collapse; font-size: 12px; font-variant-numeric: tabular-nums; }
table th { background: var(--black); color: #FFF; padding: 9px 10px; text-align: left; font-size: 10px; text-transform: uppercase; letter-spacing: 0.06em; font-weight: 700; border: 1px solid var(--slate-200); }
table th.num, table td.num { text-align: right; }
table td { padding: 7px 10px; border: 1px solid var(--slate-200); font-weight: 500; }
table tbody tr:nth-child(even) td { background: var(--slate-50); }
table tbody tr.total td { background: var(--navy); color: #FFF; font-weight: 700; border-color: var(--navy); }
table.trend { font-size: 11px; } table.trend td, table.trend th { padding: 6px 8px; }
table.trend td.year-total, table.trend th.year-total { background: var(--navy); color: #FFF; font-weight: 700; }
.neg { color: var(--neg); } .pos { color: var(--pos); }
tr.main-cat { cursor: pointer; transition: background 0.12s; }
tr.main-cat:hover td { background: var(--soft-tint) !important; }
tr.main-cat td:first-child { font-weight: 700; color: var(--navy); position: relative; padding-right: 28px; }
tr.main-cat .chevron { position: absolute; right: 8px; top: 50%; transform: translateY(-50%); color: var(--slate-500); font-size: 10px; transition: transform 0.15s; display: inline-block; }
tr.main-cat.open .chevron { transform: translateY(-50%) rotate(90deg); color: var(--navy); }
tr.sub-cat { display: none; background: var(--soft-tint); }
tr.sub-cat.show { display: table-row; }
tr.sub-cat td { color: var(--slate-600); padding-left: 28px; font-size: 11px; font-style: italic; border-top: 1px dashed var(--slate-200); }
tr.sub-cat td:first-child::before { content: '↳  '; color: var(--slate-300); font-style: normal; }
.chart-wrap { background: #FFF; border: 1px solid var(--slate-200); padding: 18px; border-radius: 6px; margin-bottom: 28px; }
.chart-wrap h4 { font-size: 11px; text-transform: uppercase; color: var(--slate-500); margin-bottom: 10px; font-weight: 700; letter-spacing: 0.08em; }
.two-col { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
@media (max-width: 900px) { .two-col { grid-template-columns: 1fr; } }
.footer { margin-top: 48px; padding-top: 20px; border-top: 1px solid var(--slate-200); color: var(--slate-500); font-size: 11px; font-style: italic; text-align: center; }
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>KSESTOR</h1>
    <h2>Monthly Sales Report Dashboard</h2>
    <div class="meta" id="meta">Loading…</div>
  </div>
  <div class="tabs">
    <button class="tab active" data-tab="exec">Executive Summary</button>
    <button class="tab" data-tab="amazon">Amazon</button>
    <button class="tab" data-tab="tiktok">TikTok Shop</button>
    <button class="tab" data-tab="shopify">Shopify & Faire</button>
    <button class="tab" data-tab="walmart">Walmart</button>
  </div>
  <div class="tab-content active" id="tab-exec">
    <div class="selectors">
      <div class="sel-group"><label>Month</label><select id="exec-month"></select></div>
      <div class="sel-group"><label>Year</label><select id="exec-year"></select></div>
      <div class="sel-group"><label>View</label><select id="exec-view"><option value="monthly">Monthly</option><option value="annual">Full Year</option></select></div>
    </div>
    <div id="exec-body"></div>
  </div>
  <div class="tab-content" id="tab-amazon">
    <div class="selectors">
      <div class="sel-group"><label>Month</label><select id="amz-month"></select></div>
      <div class="sel-group"><label>Year</label><select id="amz-year"></select></div>
      <div class="sel-group"><label>Marketplace</label><select id="amz-mp"><option>Amazon US</option><option>Amazon CA</option><option>Amazon UK</option></select></div>
    </div>
    <div id="amz-body"></div>
  </div>
  <div class="tab-content" id="tab-tiktok">
    <div class="selectors">
      <div class="sel-group"><label>Month</label><select id="tt-month"></select></div>
      <div class="sel-group"><label>Year</label><select id="tt-year"></select></div>
    </div>
    <div id="tt-body"></div>
  </div>
  <div class="tab-content" id="tab-shopify">
    <div class="selectors">
      <div class="sel-group"><label>Month</label><select id="sh-month"></select></div>
      <div class="sel-group"><label>Year</label><select id="sh-year"></select></div>
    </div>
    <div id="sh-body"></div>
  </div>
  <div class="tab-content" id="tab-walmart">
    <div class="selectors">
      <div class="sel-group"><label>Month</label><select id="wm-month"></select></div>
      <div class="sel-group"><label>Year</label><select id="wm-year"></select></div>
    </div>
    <div id="wm-body"></div>
  </div>
  <div class="footer">KSESTOR · Prepared for Kseniia · Generated from Monthly Report raw data</div>
</div>
<script>
const DATA = __DATA_PLACEHOLDER__;
const MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
const fmtMoney = v => v == null || v === '' || v === '-' || isNaN(v) ? '—' : '$' + Number(v).toLocaleString('en-US', {minimumFractionDigits: 2, maximumFractionDigits: 2});
const fmtMoneyCompact = v => v == null || v === '' || v === '-' || isNaN(v) ? '—' : '$' + Number(v).toLocaleString('en-US', {minimumFractionDigits: 0, maximumFractionDigits: 0});
const fmtPct = v => v == null || v === '' || v === '-' || isNaN(v) ? '—' : (Number(v)*100).toFixed(2) + '%';
const fmtNum = v => v == null ? '—' : Number(v).toLocaleString('en-US');
const num = v => typeof v === 'number' ? v : parseFloat(v) || 0;
function initSelect(el, items) { el.innerHTML = items.map(i => `<option>${i}</option>`).join(''); }
['exec-month','amz-month','tt-month','sh-month','wm-month'].forEach(id => initSelect(document.getElementById(id), MONTHS));
const years = [...new Set(DATA.sales.map(r=>r.year).filter(Boolean))].sort();
['exec-year','amz-year','tt-year','sh-year','wm-year'].forEach(id => initSelect(document.getElementById(id), years));
// Default to latest month/year in data
const latest = DATA.sales.length ? DATA.sales[DATA.sales.length - 1] : null;
const defaultMonth = latest && latest.month ? latest.month : 'Mar';
const defaultYear = latest && latest.year ? latest.year : 2026;
document.querySelectorAll('#exec-month,#amz-month,#tt-month,#sh-month,#wm-month').forEach(s => s.value = defaultMonth);
document.querySelectorAll('#exec-year,#amz-year,#tt-year,#sh-year,#wm-year').forEach(s => s.value = defaultYear);
document.getElementById('meta').textContent = `Generated ${DATA.generated}  ·  Amazon US · CA · UK  ·  TikTok Shop · Shopify & Faire · Walmart`;
document.querySelectorAll('.tab').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.tab').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
    document.getElementById('tab-' + btn.dataset.tab).classList.add('active');
  });
});
function filterSales(month, year, mp) { return DATA.sales.filter(r => r.month === month && r.year == year && (!mp || r.mp === mp)); }
function filterExp(month, year, mp) { return DATA.expenses.filter(r => r.month === month && r.year == year && (!mp || r.mp === mp)); }
function sumField(rows, field) { return rows.reduce((s,r) => s + (num(r[field]) || 0), 0); }
function summaryRow(rows) { return rows.find(r => r.units != null && r.units !== '' && r.units !== 0) || rows[0] || {}; }
function kpi(label, value, cls='') { return `<div class="kpi"><div class="label">${label}</div><div class="value ${cls}">${value}</div></div>`; }
function kpi$(label, v) { const cls = (typeof v === 'number' && v < 0) ? 'neg' : ''; return kpi(label, fmtMoney(v), cls); }
function kpiPct(label, v) { const cls = (typeof v === 'number' && v < 0) ? 'neg' : ''; return kpi(label, fmtPct(v), cls); }

function renderExec() {
  const m = document.getElementById('exec-month').value;
  const y = +document.getElementById('exec-year').value;
  const view = document.getElementById('exec-view').value;
  const body = document.getElementById('exec-body');
  const mps = ['Amazon US','Amazon CA','Amazon UK','Tiktok','Shopify','Walmart'];
  const mpDisplay = {'Amazon US':'Amazon US','Amazon CA':'Amazon CA','Amazon UK':'Amazon UK','Tiktok':'TikTok Shop','Shopify':'Shopify & Faire','Walmart':'Walmart'};
  if (view === 'monthly') {
    let totSales=0, totExp=0;
    const rows = mps.map(mp => {
      const s = sumField(filterSales(m,y,mp), 'sales');
      const e = sumField(filterExp(m,y,mp), 'amount');
      const np = s + e; const margin = s ? np/s : 0;
      totSales += s; totExp += e;
      return {mp: mpDisplay[mp], s, e, np, margin};
    });
    rows.forEach(r => r.pct = totSales ? r.s/totSales : 0);
    const totNP = totSales + totExp; const totMargin = totSales ? totNP/totSales : 0;
    body.innerHTML = `<div class="kpi-grid">${kpi$('Total Sales', totSales)}${kpi$('Total Expenses', totExp)}${kpi$('Net Profit', totNP)}${kpiPct('Margin', totMargin)}</div>
      <div class="section"><div class="section-head"><h3>${m} ${y} · All Marketplaces</h3></div>
      <table><thead><tr><th>Marketplace</th><th class="num">Sales (USD)</th><th class="num">Expenses (USD)</th><th class="num">Net Profit</th><th class="num">Margin</th><th class="num">Sales %</th></tr></thead>
      <tbody>${rows.map(r => `<tr><td><strong>${r.mp}</strong></td><td class="num">${fmtMoney(r.s)}</td><td class="num neg">${fmtMoney(r.e)}</td><td class="num ${r.np<0?'neg':'pos'}">${fmtMoney(r.np)}</td><td class="num ${r.margin<0?'neg':''}">${fmtPct(r.margin)}</td><td class="num">${fmtPct(r.pct)}</td></tr>`).join('')}
      <tr class="total"><td>TOTAL</td><td class="num">${fmtMoney(totSales)}</td><td class="num">${fmtMoney(totExp)}</td><td class="num">${fmtMoney(totNP)}</td><td class="num">${fmtPct(totMargin)}</td><td class="num">100.00%</td></tr></tbody></table></div>
      <div class="chart-wrap"><h4>Channel Mix · ${m} ${y}</h4><canvas id="exec-chart" height="80"></canvas></div>`;
    setTimeout(() => {
      new Chart(document.getElementById('exec-chart'), {
        type: 'bar',
        data: {labels: rows.map(r => r.mp), datasets: [
          {label: 'Sales', data: rows.map(r => r.s), backgroundColor: '#272757'},
          {label: 'Net Profit', data: rows.map(r => r.np), backgroundColor: '#E5E7F3', borderColor: '#272757', borderWidth: 2}
        ]},
        options: {responsive:true, plugins:{legend:{position:'top'}}, scales:{y:{beginAtZero:true}}}
      });
    }, 50);
  } else {
    const salesByMonth = MONTHS.map(mn => sumField(filterSales(mn, y), 'sales'));
    const expByMonth = MONTHS.map(mn => sumField(filterExp(mn, y), 'amount'));
    const npByMonth = salesByMonth.map((s,i) => s + expByMonth[i]);
    const totalAnnualSales = salesByMonth.reduce((a,b)=>a+b,0);
    const totalAnnualExp = expByMonth.reduce((a,b)=>a+b,0);
    const totalAnnualNP = totalAnnualSales + totalAnnualExp;
    let tableHtml = `<div class="kpi-grid">${kpi$(`${y} Sales`, totalAnnualSales)}${kpi$(`${y} Expenses`, totalAnnualExp)}${kpi$(`${y} Net Profit`, totalAnnualNP)}${kpiPct(`${y} Margin`, totalAnnualSales ? totalAnnualNP/totalAnnualSales : 0)}</div>
      <div class="section"><div class="section-head"><h3>Year ${y} · Monthly Trend · All Marketplaces</h3></div>
      <table class="trend"><thead><tr><th>Marketplace</th>${MONTHS.map(mn=>`<th class="num">${mn}</th>`).join('')}<th class="num year-total">Year Total</th></tr></thead><tbody>`;
    mps.forEach(mp => {
      const monthlyS = MONTHS.map(mn => sumField(filterSales(mn,y,mp), 'sales'));
      const yTotal = monthlyS.reduce((a,b)=>a+b,0);
      tableHtml += `<tr><td><strong>${mpDisplay[mp]}</strong></td>${monthlyS.map(v=>`<td class="num">${fmtMoneyCompact(v)}</td>`).join('')}<td class="num year-total">${fmtMoneyCompact(yTotal)}</td></tr>`;
    });
    tableHtml += `<tr class="total"><td>TOTAL SALES</td>${salesByMonth.map(v=>`<td class="num">${fmtMoneyCompact(v)}</td>`).join('')}<td class="num">${fmtMoneyCompact(totalAnnualSales)}</td></tr>`;
    tableHtml += `<tr class="total"><td>NET PROFIT</td>${npByMonth.map(v=>`<td class="num">${fmtMoneyCompact(v)}</td>`).join('')}<td class="num">${fmtMoneyCompact(totalAnnualNP)}</td></tr></tbody></table></div>`;
    tableHtml += `<div class="chart-wrap"><h4>Monthly Sales Trend · ${y}</h4><canvas id="exec-chart" height="80"></canvas></div>`;
    body.innerHTML = tableHtml;
    setTimeout(() => {
      new Chart(document.getElementById('exec-chart'), {
        type: 'line',
        data: {labels: MONTHS, datasets: [
          {label: 'Total Sales', data: salesByMonth, borderColor: '#272757', backgroundColor: 'rgba(39,39,87,0.1)', tension: 0.3, fill: true},
          {label: 'Net Profit', data: npByMonth, borderColor: '#059669', backgroundColor: 'rgba(5,150,105,0.1)', tension: 0.3, fill: false}
        ]},
        options: {responsive:true, plugins:{legend:{position:'top'}}}
      });
    }, 50);
  }
}
['exec-month','exec-year','exec-view'].forEach(id => document.getElementById(id).addEventListener('change', renderExec));
renderExec();

function renderMpDashboard(bodyEl, m, y, mp, kpis, trendFields, trendYearKey) {
  const srows = filterSales(m,y,mp);
  const erows = filterExp(m,y,mp);
  const summary = summaryRow(srows);
  const totSales = sumField(srows, 'sales');
  const totExp = sumField(erows, 'amount');
  const currentTrendYear = window[trendYearKey] || y;
  let kpiHtml = '<div class="kpi-grid">';
  kpis.forEach(([label, field, type]) => {
    let v;
    if (field === '_sales') v = totSales;
    else if (field === '_expenses') v = totExp;
    else v = summary[field];
    if (type === 'pct') kpiHtml += kpiPct(label, v);
    else if (type === 'num') kpiHtml += kpi(label, v == null ? '—' : fmtNum(v) + ' units');
    else kpiHtml += kpi$(label, v);
  });
  kpiHtml += '</div>';
  let trendHtml = `<div class="section">
    <div class="section-head"><h3>Monthly Trend</h3>
      <div class="inline-sel"><label>Trend Year:</label><select id="${trendYearKey}-sel">${years.map(yr=>`<option${yr==currentTrendYear?' selected':''}>${yr}</option>`).join('')}</select></div>
    </div>
    <table class="trend"><thead><tr><th>Metric</th>${MONTHS.map(mn=>`<th class="num">${mn}</th>`).join('')}<th class="num year-total">Year Total</th></tr></thead><tbody>`;
  trendFields.forEach(([label, field, fmt]) => {
    const vals = MONTHS.map(mn => {
      if (field === '_sales') return sumField(filterSales(mn,currentTrendYear,mp),'sales');
      if (field === '_expenses') return sumField(filterExp(mn,currentTrendYear,mp),'amount');
      const sr = summaryRow(filterSales(mn,currentTrendYear,mp));
      return sr[field];
    });
    let total;
    if (fmt === 'pct') {
      const nums = vals.filter(v => typeof v === 'number');
      total = nums.length ? nums.reduce((a,b)=>a+b,0) / nums.length : null;
    } else {
      total = vals.reduce((a,b) => a + (typeof b === 'number' ? b : 0), 0);
    }
    trendHtml += `<tr><td><strong>${label}</strong></td>`;
    vals.forEach(v => {
      const formatted = fmt==='pct' ? fmtPct(v) : fmt==='num' ? fmtNum(v) : fmtMoneyCompact(v);
      const cls = (typeof v === 'number' && v < 0) ? 'neg' : '';
      trendHtml += `<td class="num ${cls}">${formatted}</td>`;
    });
    const totalFmt = fmt==='pct' ? fmtPct(total) : fmt==='num' ? fmtNum(total) : fmtMoneyCompact(total);
    trendHtml += `<td class="num year-total">${totalFmt}</td></tr>`;
  });
  trendHtml += '</tbody></table></div>';
  const mainCatTotals = {};
  erows.forEach(r => {
    const k = (r.category || '').trim() || 'Other';
    mainCatTotals[k] = (mainCatTotals[k] || 0) + (r.amount || 0);
  });
  const subByMain = {};
  DATA.sub_expenses.filter(r => r.month === m && r.year == y && r.mp === mp).forEach(r => {
    const main = r.main || 'Other';
    if (!subByMain[main]) subByMain[main] = [];
    subByMain[main].push(r);
  });
  const subCatSales = {};
  srows.forEach(r => {
    const k = (r.sub_cat || '').trim() || 'Other';
    subCatSales[k] = (subCatSales[k] || 0) + (r.sales || 0);
  });
  const mainCatEntries = Object.entries(mainCatTotals).filter(([k,v])=>v!==0).sort((a,b)=>a[1]-b[1]);
  const salesEntries = Object.entries(subCatSales).filter(([k,v])=>v!==0).sort((a,b)=>b[1]-a[1]);
  let breakdownHtml = '<div class="two-col">';
  breakdownHtml += `<div class="section"><div class="section-head"><h3>Expense Breakdown · ${m} ${y}</h3></div>
    <table><thead><tr><th>Category</th><th class="num">Amount</th><th class="num">%</th></tr></thead><tbody>`;
  mainCatEntries.forEach(([cat, amt], idx) => {
    const pct = totExp ? amt/totExp : 0;
    const subs = subByMain[cat] || [];
    const hasSubs = subs.length > 0;
    const rowId = `${trendYearKey}-mc-${idx}`;
    breakdownHtml += `<tr class="main-cat" data-main="${rowId}" ${hasSubs?'onclick="toggleSubs(\''+rowId+'\')"':''}>
      <td>${cat}${hasSubs?'<span class="chevron">▶</span>':''}</td>
      <td class="num neg">${fmtMoney(amt)}</td><td class="num">${fmtPct(pct)}</td></tr>`;
    subs.forEach(s => {
      const subPct = totExp ? s.amount/totExp : 0;
      breakdownHtml += `<tr class="sub-cat" data-parent="${rowId}"><td>${s.sub}</td><td class="num neg">${fmtMoney(s.amount)}</td><td class="num">${fmtPct(subPct)}</td></tr>`;
    });
  });
  breakdownHtml += `<tr class="total"><td>TOTAL</td><td class="num">${fmtMoney(totExp)}</td><td class="num">100.00%</td></tr></tbody></table></div>`;
  breakdownHtml += `<div class="section"><div class="section-head"><h3>Sales Breakdown · ${m} ${y}</h3></div>
    <table><thead><tr><th>Sub Category</th><th class="num">Amount</th><th class="num">%</th></tr></thead><tbody>`;
  salesEntries.forEach(([cat, amt]) => {
    const pct = totSales ? amt/totSales : 0;
    breakdownHtml += `<tr><td>${cat}</td><td class="num">${fmtMoney(amt)}</td><td class="num">${fmtPct(pct)}</td></tr>`;
  });
  breakdownHtml += `<tr class="total"><td>TOTAL</td><td class="num">${fmtMoney(totSales)}</td><td class="num">100.00%</td></tr></tbody></table></div></div>`;
  bodyEl.innerHTML = kpiHtml + trendHtml + breakdownHtml;
  const ts = document.getElementById(`${trendYearKey}-sel`);
  if (ts) {
    ts.addEventListener('change', () => {
      window[trendYearKey] = +ts.value;
      if (trendYearKey === 'amzTY') renderAmazon();
      else if (trendYearKey === 'ttTY') renderTikTok();
      else if (trendYearKey === 'shTY') renderShopify();
      else if (trendYearKey === 'wmTY') renderWalmart();
    });
  }
}
window.toggleSubs = function(rowId) {
  const mainRow = document.querySelector(`tr.main-cat[data-main="${rowId}"]`);
  const subRows = document.querySelectorAll(`tr.sub-cat[data-parent="${rowId}"]`);
  if (!mainRow) return;
  const isOpen = mainRow.classList.toggle('open');
  subRows.forEach(r => r.classList.toggle('show', isOpen));
};

const AMZ_KPIS = [['SALES','_sales','$'],['EXPENSES','_expenses','$'],['UNIT SALES','units','num'],['VAT','vat','$'],['NET PROFIT','net_profit','$'],['TACOS','tacos','pct'],['% REFUNDS','refunds','pct'],['SELLABLE RETURNS','sellable_returns','pct'],['MARGIN','margin','pct'],['ROI','roi','pct']];
const AMZ_TREND = [['Sales (USD)','_sales','$'],['Units','units','num'],['Net Profit','net_profit','$'],['TACOS','tacos','pct'],['Margin','margin','pct']];
function renderAmazon() { renderMpDashboard(document.getElementById('amz-body'), document.getElementById('amz-month').value, +document.getElementById('amz-year').value, document.getElementById('amz-mp').value, AMZ_KPIS, AMZ_TREND, 'amzTY'); }
['amz-month','amz-year','amz-mp'].forEach(id => document.getElementById(id).addEventListener('change', renderAmazon));
renderAmazon();

const TT_KPIS = [['NET SALES','tt_net_sales','$'],['UNIT SALES','units','num'],['TIKTOK SHIPPING','tt_shipping','$'],['MCF SHIPPING','tt_mcf','$'],['3PL LABEL COST','tt_3pl','$'],['FEES','tt_fees','$'],['REFERRAL FEE','tt_referral','$'],['AFFILIATE COMM.','tt_affiliate','$'],['ADJUSTMENTS','tt_adj','$'],['COGS','tt_cogs','$'],['ADVERTISING','tt_ads','$'],['NET PROFIT','tt_np','$'],['TACOS','tacos','pct'],['% REFUNDS','refunds','pct'],['MARGIN','tt_margin','pct']];
const TT_TREND = [['Net Sales (USD)','tt_net_sales','$'],['Units','units','num'],['COGS','tt_cogs','$'],['Ad Cost','tt_ads','$'],['Net Profit','tt_np','$'],['TACOS','tacos','pct'],['Margin','tt_margin','pct']];
function renderTikTok() { renderMpDashboard(document.getElementById('tt-body'), document.getElementById('tt-month').value, +document.getElementById('tt-year').value, 'Tiktok', TT_KPIS, TT_TREND, 'ttTY'); }
['tt-month','tt-year'].forEach(id => document.getElementById(id).addEventListener('change', renderTikTok));
renderTikTok();

const SH_KPIS = [['SALES','_sales','$'],['UNIT SALES','units','num'],['ADVERTISING','sh_ads','$'],['SHIPPING','sh_ship','$'],['MCF SHIPPING','sh_mcf','$'],['3PL LABEL COST','sh_3pl','$'],['PAYMENT FEES','sh_pay','$'],['COGS','sh_cogs','$'],['TAX','sh_tax','$'],['GROSS PROFIT','sh_gp','$'],['NET PROFIT','sh_np','$'],['TACOS','sh_tacos','pct'],['% REFUNDS','sh_refunds','pct'],['MARGIN','sh_margin','pct'],['ROI','sh_roi','pct']];
const SH_TREND = [['Sales (USD)','_sales','$'],['Units','units','num'],['Ad Cost','sh_ads','$'],['COGS','sh_cogs','$'],['Gross Profit','sh_gp','$'],['Net Profit','sh_np','$'],['TACOS','sh_tacos','pct'],['Margin','sh_margin','pct']];
function renderShopify() { renderMpDashboard(document.getElementById('sh-body'), document.getElementById('sh-month').value, +document.getElementById('sh-year').value, 'Shopify', SH_KPIS, SH_TREND, 'shTY'); }
['sh-month','sh-year'].forEach(id => document.getElementById(id).addEventListener('change', renderShopify));
renderShopify();

const WM_KPIS = [['SALES','_sales','$'],['UNIT SALES','units','num'],['ADVERTISING','wm_ads','$'],['WFS SHIPPING','wm_wfs','$'],['MCF SHIPPING','wm_mcf','$'],['3PL LABEL COST','wm_3pl','$'],['FEES (Commission)','wm_fee_com','$'],['FEES (Storage)','wm_fee_store','$'],['OTHER FEES','wm_other','$'],['COGS','wm_cogs','$'],['NET PROFIT','wm_np','$'],['TACOS','tacos','pct'],['% REFUNDS','refunds','pct'],['MARGIN','margin','pct']];
const WM_TREND = [['Sales (USD)','_sales','$'],['Units','units','num'],['Ad Cost','wm_ads','$'],['COGS','wm_cogs','$'],['Net Profit','wm_np','$'],['TACOS','tacos','pct'],['Margin','margin','pct']];
function renderWalmart() { renderMpDashboard(document.getElementById('wm-body'), document.getElementById('wm-month').value, +document.getElementById('wm-year').value, 'Walmart', WM_KPIS, WM_TREND, 'wmTY'); }
['wm-month','wm-year'].forEach(id => document.getElementById(id).addEventListener('change', renderWalmart));
renderWalmart();
</script>
</body>
</html>
'''

html = HTML_TEMPLATE.replace("__DATA_PLACEHOLDER__", data_js)

with open(HTML_FILE, "w", encoding="utf-8") as f:
    f.write(html)

print(f"\nDashboard regenerated: {HTML_FILE.name}")
print(f"Size: {os.path.getsize(HTML_FILE)/1024:.1f} KB")
print("\nNext step:")
print("  1. Open the HTML file locally to preview (double-click it).")
print("  2. If it looks right, copy it to your GitHub folder as 'index.html' and push.")
